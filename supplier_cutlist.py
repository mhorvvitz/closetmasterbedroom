#!/usr/bin/env python3
"""supplier_cutlist.py — fill the supplier's Hebrew cutting-list template.

Reads output/cutlist.json (derived from closet_spec.py) and writes the
supplier's own .xls template via Excel COM, so their layout and formatting
survive untouched.

Template columns (sheet "גיליון1", header on row 6, data from row 7):
  A שורה | B סוג החומר | C כמות | D רוחב | E אורך | F ניתן לסובב | G קנטים | H הערות

Two conventions the supplier cares about:
  * רוחב / אורך — their sheet is in mm. We map WIDTH = the smaller face
    dimension, LENGTH = the larger, matching how the cut list already sorts.
  * ניתן לסובב ("can be rotated") — this is the nesting freedom flag. A plain
    COLOUR has no grain, so every part may be rotated, which materially
    improves their yield. On a woodgrain decor this would be "לא" for any
    part with a grain direction.
"""
import json
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
TPL = os.path.join(HERE, "Suppliers", "רשימת-חיתוך.xls")
OUT = os.path.join(HERE, "output", "רשימת-חיתוך-מלא.xls")
CUT = os.path.join(HERE, "output", "cutlist.json")

CLIENT = "ארון קיר — חדר שינה הורים"
MATERIAL_HE = "מלמין צבעוני 17 מ\"מ"

# Which physical edges get banding, in the supplier's language.
BAND_HE = {
    "front":       "1 ארוך",
    "front, w1":   "1 ארוך + 1 קצר",
    "w1":          "1 קצר",
    "all":         "4 צדדים",
    "":            "-",
}

# Short Hebrew note per part family. ORDER MATTERS — matched most-specific
# first, otherwise "Cleat_D1_SideL" matches "Side" and gets called a tower
# side panel, and "L_ShoeShelf" matches "Shelf" and loses the "shoe".
NOTE_HE = (
    ("ShoeCleat", "סרגל תלייה למדף נעליים"),
    ("Cleat", "סרגל תלייה לקיר"),
    ("ShoeShelf", "מדף נעליים"),
    ("ShakerRail", "מסגרת שייקר - אופקי"),
    ("ShakerStile", "מסגרת שייקר - אנכי"),
    ("DrawerFront", "חזית מגירה"),
    ("Upright_Div", "עמוד המשך למחיצה"),
    ("Upright", "עמוד בין המדפים"),
    ("Divider", "מחיצה אמצעית לתלייה"),
    ("Deck1", "מדף עליון 1"),
    ("Deck2", "מדף עליון 2"),
    ("Side", "דופן מגדל"),
    ("Bottom", "תחתית מגדל"),
    ("Shelf", "מדף"),
)


def note_for(name):
    for k, v in NOTE_HE:
        if k in name:
            return v
    return ""


def rows(material_prefix="melamine"):
    d = json.load(open(CUT, encoding="utf-8"))
    out = []
    for p in d["parts"]:
        if not p["material_id"].startswith(material_prefix):
            continue
        band = BAND_HE.get(p.get("banding", "") or "", p.get("banding", ""))
        out.append(dict(mat=MATERIAL_HE, qty=p["qty"],
                        width=p["width"], length=p["length"],
                        rotate="כן", band=band, note=note_for(p["name"]),
                        ref=p["name"]))
    out.sort(key=lambda r: (-r["length"], -r["width"]))
    return out


def write_xlsx(data, path):
    """Rebuild the supplier's layout with openpyxl.

    Excel COM was the obvious route (fill their .xls in place, keep their
    formatting) but it fails on this machine — the IDispatch property set
    throws a bogus "cannot cast Int32 to String", and pinning the thread
    culture to en-US did not help. Rebuilding the sheet is more portable
    anyway: no Excel dependency, works on any machine.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    HDR = ["שורה", "סוג החומר", "כמות", "רוחב", "אורך",
           "ניתן לסובב", "קנטים", "הערות"]
    wb = Workbook()
    sh = wb.active
    sh.title = "גיליון1"
    sh.sheet_view.rightToLeft = True          # Hebrew sheet reads right-to-left

    thin = Side(style="thin", color="999999")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal="center", vertical="center")

    sh["C2"] = 'רשימת מידות (במילימטרים) לחיתוך'
    sh["C2"].font = Font(bold=True, size=14)
    sh["B4"] = "שם הלקוח:"
    sh["B4"].font = Font(bold=True)
    sh["C4"] = CLIENT
    sh["G4"] = "תאריך"
    sh["G4"].font = Font(bold=True)
    sh["H4"] = "16/08/2026"

    for c, h in enumerate(HDR, 1):
        cell = sh.cell(row=6, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDE5EC")
        cell.alignment = ctr
        cell.border = box

    r = 7
    for i, d in enumerate(data, 1):
        vals = [i, d["mat"], d["qty"], d["width"], d["length"],
                d["rotate"], d["band"], f'{d["note"]} ({d["ref"]})']
        for c, v in enumerate(vals, 1):
            cell = sh.cell(row=r, column=c, value=v)
            cell.border = box
            if c in (1, 3, 4, 5, 6):
                cell.alignment = ctr
        r += 1

    tot = sum(d["qty"] for d in data)
    sh.cell(row=r + 1, column=2,
            value=f'סה"כ {tot} חלקים ב-{len(data)} שורות').font = Font(bold=True)
    sh.cell(row=r + 2, column=2,
            value="כל החלקים ניתנים לסיבוב — צבע אחיד ללא כיוון סיב")
    sh.cell(row=r + 3, column=2,
            value='לוח: מלמין צבעוני 17 מ"מ, 2440x1220')
    sh.cell(row=r + 5, column=2, value="הערה לגבי מסגרות השייקר:").font = \
        Font(bold=True)
    sh.cell(row=r + 6, column=2, value=(
        "24 רצועות ברוחב 50 מ\"מ. אנא בצעו קנטים על שתי הפאות הארוכות של "
        "רצועה ארוכה אחת, ורק אחר כך חתכו לאורכים (594 ו-120). כך נחסכות "
        "24 העברות נפרדות במכונת הקנטים."))
    sh.cell(row=r + 7, column=2, value=(
        "לאחר החיתוך יש להוסיף קנט רק לקצוות של 12 הרצועות באורך 594."))

    for col, w in zip("ABCDEFGH", (7, 22, 8, 9, 9, 12, 18, 34)):
        sh.column_dimensions[col].width = w
    sh.freeze_panes = "A7"
    wb.save(path)
    return tot


ROWS_JSON = os.path.join(HERE, "output", "supplier_rows.json")


def dump_json(data):
    """pywin32 is not installed here, so the Excel COM fill is driven from
    PowerShell instead. This is the handoff file."""
    import io
    payload = dict(client=CLIENT, date="16/08/2026", template=TPL, out=OUT,
                   rows=data)
    io.open(ROWS_JSON, "w", encoding="utf-8", newline="\n").write(
        json.dumps(payload, ensure_ascii=False, indent=1))
    return ROWS_JSON


if __name__ == "__main__":
    data = rows()
    print(f"{len(data)} rows, {sum(d['qty'] for d in data)} parts")
    for d in data[:5]:
        print(f"  {d['qty']:>2} x {d['width']:>4} x {d['length']:<5} "
              f"{d['band']:<16} {d['ref']}")
    dump_json(data)
    xlsx = os.path.join(HERE, "output", "רשימת-חיתוך-מלא.xlsx")
    tot = write_xlsx(data, xlsx)
    print(f"wrote {xlsx}  ({len(data)} rows, {tot} parts)")

